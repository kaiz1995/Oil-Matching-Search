#!/usr/bin/env python3
"""
Unified build pipeline: Excel → JSON → HTML
1. Reads 竞品产品对照表.xlsx as PRIMARY source (user edits this file)
2. Reads product_data.json (detailed Runmei mappings, as supplement)
3. Falls back to competitor_final.json if Excel unavailable
4. Smart matching to fill Runmei where possible
5. Outputs: product_data_merged.json + lubricant_product_matching_tool.html
"""
import json, re, sys, os, base64, argparse, shutil, subprocess, urllib.parse
sys.stdout.reconfigure(encoding='utf-8')

parser = argparse.ArgumentParser(description='润滑油数据构建管线')
parser.add_argument('--release', action='store_true', help='构建 APK 并创建 GitHub Release')
parser.add_argument('--skip-git', action='store_true', help='跳过 git 提交和推送')
parser.add_argument('--skip-apk', action='store_true', help='跳过 APK 构建（仅测试 --release 其他步骤）')
args = parser.parse_args()

BASE_DIR = r'D:\HuaweiMoveData\Users\张大脸小太阳\Documents\cc workspace\润滑产品查询工具'
CATALOG = f'{BASE_DIR}/competitor_final.json'
EXCEL_PRIMARY = f'{BASE_DIR}/竞品产品对照表.xlsx'
EXCEL_FILLED = f'{BASE_DIR}/竞品产品对照表_已填充.xlsx'

# ============================================================
# STEP 1: Load data sources
# ============================================================

# Load mapped data from data.json (remap field names to match script expectations)
with open(f'{BASE_DIR}/data.json', 'r', encoding='utf-8') as f:
    mapped_json = json.load(f)
mapped_products = mapped_json.get('products', mapped_json if isinstance(mapped_json, list) else [])
mappings = []
for m in mapped_products:
    if not isinstance(m, dict):
        continue
    m['competitorProduct'] = m.get('product_cn', m.get('competitorProduct', ''))
    m['competitorBrand'] = m.get('brand', m.get('competitorBrand', ''))
    m['runmeiProduct'] = m.get('runmei', m.get('runmeiProduct', ''))
    m['competitorProductEN'] = m.get('product_en', m.get('competitorProductEN', ''))
    m['sourceSheet'] = m.get('source', m.get('sourceSheet', ''))
    if m.get('competitorProduct'):
        mappings.append(m)
print(f"Mapped products: {len(mappings)}")

# Try reading from Excel first (user's primary editing surface)
def read_catalog_from_excel(excel_path):
    """Read product catalog from Excel file. Returns list of dicts or None."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        # Use the correct sheet: '竞品产品对照表' (has the data), not the active sheet
        ws = wb['竞品产品对照表'] if '竞品产品对照表' in wb.sheetnames else wb.active
        headers = [str(ws.cell(1, c).value or '') for c in range(1, ws.max_column + 1)]
        # Map headers to field names
        col_map = {}
        for i, h in enumerate(headers):
            h_lower = h.lower()
            if '品牌' in h:
                col_map['brand'] = i
            elif '类型' in h or '产品类型' in h:
                col_map['category'] = i
            elif '竞品产品名称' in h or '产品名称' in h or h == '名称':
                col_map['product_cn'] = i
            elif '英文' in h:
                col_map['product_en'] = i
            elif '粘度' in h or 'vg' in h.lower():
                col_map['viscosity'] = i
            elif '润美' in h:
                col_map['runmei'] = i
            elif '来源' in h or '数据来源' in h:
                col_map['source'] = i

        products = []
        for row in range(2, ws.max_row + 1):
            def cell(col_name):
                idx = col_map.get(col_name)
                if idx is None:
                    return ''
                val = ws.cell(row, idx + 1).value
                return str(val).strip() if val else ''

            cn = cell('product_cn')
            brand = cell('brand')
            if not cn or not brand:
                continue  # Skip empty rows

            # Skip header-like rows
            if cn in ('竞品产品名称', '产品名称', '品牌'):
                continue

            products.append({
                'brand': brand,
                'category': cell('category'),
                'product_cn': cn,
                'product_en': cell('product_en'),
                'viscosity': cell('viscosity'),
                'runmei': cell('runmei'),
                'source': cell('source'),
            })
        wb.close()
        return products if products else None
    except Exception as e:
        print(f"  [WARN] Cannot read Excel: {e}")
        return None

def auto_field_key(header):
    """Auto-generate a field key from header text for unknown columns."""
    import re
    key = re.sub(r'[^\w一-鿿]+', '_', header).strip('_')
    if not key:
        key = 'extra'
    return key

def parse_runmei_sheet(excel_path):
    """Parse '润美产品目录' sheet into structured products + param config.

    Dynamically detects ALL columns from row 1 headers.
    Returns (products_list, param_config_list).
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet_name = next((s for s in ['润美产品目录', '润美产品查询'] if s in wb.sheetnames), None)
        if not sheet_name:
            wb.close()
            return None, []
        ws = wb[sheet_name]

        # Read all column headers from row 1
        headers = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(1, c).value
            headers.append(str(val).strip() if val else '')

        # Keyword -> field_key mapping for known columns
        HEADER_MAP = {
            '技术系列': 'series',
            '产品牌号': 'product_cn',
            '包装规格': 'packaging',
            '应用行业': 'industry',
            '主要应用部位': 'application_site',
            '产品特性': 'features',      # matches "对润滑的要求或产品特性"
            '备注': 'notes',
            '外观': 'appearance',
            '运动粘度40': 'kv40',
            '运动粘度100': 'kv100',
            '粘度指数': 'vi',
            '倾点': 'pour_point',
            '闪点': 'flash_point',
            '铜片腐蚀': 'copper_corrosion',
            '锈蚀 A': 'rust_a',
            '锈蚀 B': 'rust_b',
            'NAS1638': 'cleanliness_nas',
            '清洁度ISO': 'cleanliness_iso',
            '磨斑直径': 'wear_scar',
            '烧结负荷': 'weld_load',
            '最大无卡咬': 'pb_load',
            '空气释放': 'air_release',
            '抗乳化性(54℃)': 'demulsibility_54',
            '抗乳化性，82': 'demulsibility_82',
            '抗乳化性(82℃)': 'demulsibility_82_detail',
            '泡沫特性': 'foam',
            'PH值': 'ph_value',
            '凝点': 'solidification_point',
            '密度': 'density',
            '酸值': 'acid_value',
            'FZG': 'fzg',
            'TDS文件': 'tds_file',
            'MSDS文件': 'msds_file',
        }

        # Fields that are NOT technical parameters (excluded from paramConfig)
        CORE_FIELDS = {'series', 'product_cn', 'packaging', 'industry',
                       'application_site', 'features', 'notes', 'appearance',
                       'tds_file', 'msds_file'}

        # Known parameter labels and test standards
        PARAM_LABELS = {
            'kv40': '运动粘度40℃(mm²/s)', 'kv100': '运动粘度100℃(mm²/s)',
            'vi': '粘度指数', 'pour_point': '倾点(℃)', 'flash_point': '闪点(℃)',
            'copper_corrosion': '铜片腐蚀(100℃,3h)',
            'rust_a': '液相锈蚀A法', 'rust_b': '液相锈蚀B法',
            'cleanliness_nas': '清洁度(NAS1638)', 'cleanliness_iso': '清洁度(ISO 4406)',
            'wear_scar': '磨斑直径(mm)', 'weld_load': '烧结负荷(N)',
            'pb_load': '最大无卡咬负荷(N)',
            'air_release': '空气释放值(50℃)(min)',
            'demulsibility_54': '抗乳化性(54℃)(min)',
            'demulsibility_82': '抗乳化性(82℃)(min)',
            'demulsibility_82_detail': '抗乳化性(82℃) 乳化液/分离水/油中水',
            'foam': '泡沫特性(mL/mL)',
            'ph_value': 'pH值',
            'solidification_point': '凝点(℃)',
            'density': '密度（15°C）(g/cm³)',
            'acid_value': '酸值(mgKOH/g)',
            'fzg': 'FZG A20/8.3/90',
        }
        PARAM_STANDARDS = {
            'kv40': 'GB/T 265', 'kv100': 'GB/T 265', 'vi': 'GB/T 1995',
            'pour_point': 'GB/T 3535', 'flash_point': 'GB/T 3536',
            'copper_corrosion': 'GB/T 5096',
            'rust_a': 'GB/T 11143', 'rust_b': 'GB/T 11143',
            'cleanliness_nas': 'NAS 1638', 'cleanliness_iso': 'ISO 4406',
            'wear_scar': 'SH/T 0189', 'weld_load': 'GB/T 3142',
            'pb_load': 'GB/T 3142',
            'air_release': 'SH/T 0308',
            'demulsibility_54': 'GB/T 7305',
            'demulsibility_82': 'GB/T 7305',
            'demulsibility_82_detail': 'GB/T 7305',
            'foam': 'GB/T 12579',
            'ph_value': 'GB/T 7304',
            'solidification_point': 'GB/T 3535',
            'density': 'GB/T 1884',
            'acid_value': 'GB/T 4945',
            'fzg': 'NB/SH/T 0306-2013',
        }

        # Build column index -> field_key mapping
        col_map = {}     # 0-based index -> field_key
        col_params = []  # ordered list of {key, label, standard} tuples

        for idx, header in enumerate(headers):
            if not header:
                continue
            # Match against known keywords
            matched_key = None
            for keyword, field_key in HEADER_MAP.items():
                if keyword in header:
                    matched_key = field_key
                    break
            if matched_key:
                col_map[idx] = matched_key
                if matched_key not in CORE_FIELDS:
                    col_params.append({
                        'key': matched_key,
                        'label': PARAM_LABELS.get(matched_key, header),
                        'standard': PARAM_STANDARDS.get(matched_key, '')
                    })
            else:
                # Unknown column — auto-generate field key
                fk = auto_field_key(header)
                col_map[idx] = fk
                col_params.append({
                    'key': fk,
                    'label': header,
                    'standard': ''
                })

        # Deduplicate col_params by key (safety net for overlapping keyword matches)
        seen_keys = set()
        unique_params = []
        for param in col_params:
            if param['key'] not in seen_keys:
                seen_keys.add(param['key'])
                unique_params.append(param)
        col_params = unique_params

        # Read data rows
        products = []
        for row in range(2, ws.max_row + 1):
            def c(idx):
                val = ws.cell(row, idx + 1).value
                return str(val).strip() if val else ''
            name = c(1)  # B: 产品牌号
            if not name:
                continue
            p = {}
            for idx, field_key in col_map.items():
                val = c(idx)
                # Always set core fields even if empty; param fields only if non-empty
                if val or field_key in CORE_FIELDS:
                    if field_key in ('tds_file', 'msds_file'):
                        if val:
                            encoded_val = urllib.parse.quote(val)
                            p[field_key] = f"https://kaiz1995.github.io/Oil-Matching-Search/pdfs/{encoded_val}"
                    else:
                        p[field_key] = val
            products.append(p)

        wb.close()
        return products if products else None, col_params

    except Exception as e:
        print(f"  [WARN] Cannot read runmei catalog: {e}")
        return None, []

# Try filled Excel first (user's primary editing surface), then original, then JSON
catalog = None
for excel_path in [EXCEL_FILLED, EXCEL_PRIMARY]:
    if os.path.exists(excel_path):
        catalog = read_catalog_from_excel(excel_path)
        if catalog:
            print(f"Catalog from Excel: {len(catalog)} products ({os.path.basename(excel_path)})")
            break

if not catalog:
    # Fallback to JSON catalog
    with open(CATALOG, 'r', encoding='utf-8') as f:
        catalog = json.load(f)
    # Filter noise entries
    NOISE_WORDS = [
        '服务流程', '服务体系的组成', '全程支持', 'LubeAssist',
        'EfficiencySupport', 'Energy克鲁勃能效', 'Maintain克鲁勃维护',
        'Monitor克鲁勃监控', 'Renew克鲁勃更新', 'OilSuite',
        '技术支持', '咨询服务', '培训服务',
    ]
    catalog = [p for p in catalog if not any(k in p.get('product_cn', '') for k in NOISE_WORDS)]
    print(f"Catalog from JSON (fallback): {len(catalog)} products")

# Load Runmei product catalog from Excel
runmei_catalog = None
runmei_param_config = []
for excel_path in [EXCEL_FILLED, EXCEL_PRIMARY]:
    if os.path.exists(excel_path):
        runmei_catalog, runmei_param_config = parse_runmei_sheet(excel_path)
        if runmei_catalog:
            print(f"Runmei catalog: {len(runmei_catalog)} products, {len(runmei_param_config)} params detected")
            break

# ============================================================
# STEP 2: Normalization & matching utilities
# ============================================================

def normalize(s):
    """Normalize for comparison"""
    s = s.lower().strip()
    s = re.sub(r'[™®\s]+', ' ', s)
    s = re.sub(r'[^\w\s一-鿿]', '', s)
    return s.strip()

def extract_viscosities(name):
    return set(int(x) for x in re.findall(r'\b(\d{2,3})\b', name))

def extract_family(name):
    """Extract product family name (brand + series, without viscosity)"""
    name = re.sub(r'\d{2,3}(?:\s*/\s*\d{2,3})*', '', name)
    name = re.sub(r'\s+', ' ', name).strip()
    return name

def brand_match(b1, b2):
    """Check if two brand strings refer to the same brand"""
    b1, b2 = normalize(b1), normalize(b2)
    for b in ['壳牌', 'shell', '美孚', 'mobil', '雪佛龙', 'chevron',
              '长城', 'sinopec', '克鲁勃', 'kluber', 'klüber',
              '嘉实多', 'castrol', '福斯', 'fuchs', '安索', 'amsoil',
              '昆仑', 'kunlun']:
        if b in b1 and b in b2:
            return True
    return b1 == b2

# Category normalization
CAT_MAP = {
    '空压机油/压缩机油': '空压机油',
    '汽轮机油/涡轮机油': '涡轮机油',
    '轴承油/循环油': '轴承油',
}

def cat_match(c1, c2):
    c1 = CAT_MAP.get(c1, c1)
    c2 = CAT_MAP.get(c2, c2)
    return c1 == c2

# ============================================================
# STEP 3: Build lookup indexes from mapped data
# ============================================================

name_index = {}       # normalized name → mapping
family_index = {}     # (family, category) → [mappings]
brand_cat_index = {}  # (brand, category) → [mappings]

for m in mappings:
    if not m.get('competitorProduct'):
        continue
    norm = normalize(m['competitorProduct'])
    family = normalize(extract_family(m['competitorProduct']))
    cat = m['category']
    brand = m['competitorBrand']
    runmei = m.get('runmeiProduct', '')

    name_index[norm] = m
    if family and runmei:
        key = (family, cat)
        if key not in family_index:
            family_index[key] = []
        family_index[key].append(m)

    bck = (brand, cat)
    if bck not in brand_cat_index:
        brand_cat_index[bck] = []
    brand_cat_index[bck].append(m)

print(f"Name index: {len(name_index)} entries")
print(f"Family index: {len(family_index)} families")
print(f"Brand-cat index: {len(brand_cat_index)} groups")

# ============================================================
# STEP 4: Smart matching function
# ============================================================

def find_runmei(catalog_entry):
    """Find best Runmei match for a competitor product from catalog"""
    cn_name = catalog_entry['product_cn']
    brand = catalog_entry['brand']
    cat = catalog_entry['category']

    norm = normalize(cn_name)

    # Step 1: Exact name match
    if norm in name_index:
        m = name_index[norm]
        return m.get('runmeiProduct', ''), m

    # Step 2: Name match without brand prefix
    for prefix in ['壳牌', '美孚', '雪佛龙', '长城', '克鲁勃', '长城牌',
                   'Shell', 'Mobil', 'Chevron', 'Klüber', 'Kluber']:
        if prefix in cn_name:
            bare = normalize(cn_name.replace(prefix, ''))
            if bare in name_index:
                m = name_index[bare]
                return m.get('runmeiProduct', ''), m
            break

    # Step 3: Family-based match (same product family, different viscosity)
    cat_brand = brand.split('/')[0].strip()  # "壳牌/Shell" → "壳牌"
    cn_family = normalize(extract_family(cn_name))
    cat_vis = set(extract_viscosities(cn_name + ' ' + catalog_entry.get('viscosity', '')))

    for (family, fcat), entries in family_index.items():
        family_parts = set(family.split())
        cn_parts = set(cn_family.split())
        common = family_parts & cn_parts
        # Allow family match across related categories
        if len(common) >= 2 and cat in (fcat, '液压油', '齿轮油', '空压机油', '涡轮机油', '汽轮机油'):
            for m in entries:
                if not m.get('runmeiProduct'):
                    continue
                m_vis = set(extract_viscosities(m['competitorProduct']))
                if not cat_vis or not m_vis or (cat_vis & m_vis):
                    return m.get('runmeiProduct', ''), m
            # Fallback: only when catalog entry has no viscosity info
            if not cat_vis:
                for m in entries:
                    if m.get('runmeiProduct'):
                        return m.get('runmeiProduct', ''), m

    # Step 4: Brand + category based fallback
    bck = (cat_brand, cat)
    if bck in brand_cat_index:
        entries = brand_cat_index[bck]
        cat_vis = extract_viscosities(catalog_entry.get('viscosity', '') + ' ' + cn_name)
        for m in entries:
            if not m.get('runmeiProduct'):
                continue
            m_vis = extract_viscosities(m.get('competitorProduct', '') + ' ' + m.get('competitorProductEN', ''))
            if cat_vis and m_vis and (set(cat_vis) & set(m_vis)):
                return m.get('runmeiProduct', ''), m
        # Fallback: only when catalog entry has no viscosity info
        if not cat_vis:
            for m in entries:
                if m.get('runmeiProduct'):
                    return m.get('runmeiProduct', ''), m

    return '', None

# ============================================================
# STEP 5: Process all catalog entries
# ============================================================

merged = []
matched_count = 0
user_filled_count = 0

for entry in catalog:
    # Check if user already filled Runmei in Excel - that takes priority
    user_runmei = entry.get('runmei', '').strip()
    if user_runmei:
        runmei = user_runmei
        mapping = None
        user_filled_count += 1
        # Try to find mapping for specs
        _, mapping = find_runmei(entry)
    else:
        runmei, mapping = '', None

    merged.append({
        'brand': entry['brand'],
        'category': entry['category'],
        'product_cn': entry['product_cn'],
        'product_en': entry.get('product_en', ''),
        'viscosity': entry.get('viscosity', ''),
        'source': entry.get('source', ''),
        'runmei': runmei,
        'application': mapping.get('application', '') if mapping else '',
        'compSpecs': mapping.get('compSpecs', {}) if mapping else {},
        'runmeiSpecs': mapping.get('runmeiSpecs', {}) if mapping else {},
    })

total_runmei = matched_count + user_filled_count
print(f"User-filled: {user_filled_count}, Auto-matched: {matched_count}, Total Runmei: {total_runmei}/{len(merged)} ({100*total_runmei/len(merged):.1f}%)")

# Add mapped products not in catalog
existing_names = set(normalize(m['product_cn']) for m in merged)
for m in mappings:
    if not m.get('competitorProduct'):
        continue
    norm = normalize(m['competitorProduct'])
    if norm not in existing_names:
        existing_names.add(norm)
        merged.append({
            'brand': m['competitorBrand'],
            'category': m['category'],
            'product_cn': m['competitorProduct'],
            'product_en': m.get('competitorProductEN', ''),
            'viscosity': '',
            'source': m.get('sourceSheet', ''),
            'runmei': '',
            'application': m.get('application', ''),
            'compSpecs': m.get('compSpecs', {}),
            'runmeiSpecs': m.get('runmeiSpecs', {}),
        })

# Deduplicate by product_cn, keeping entries with runmei
deduped = {}
for m in merged:
    key = normalize(m['product_cn'])
    if key not in deduped or (m['runmei'] and not deduped[key]['runmei']):
        deduped[key] = m
merged = list(deduped.values())
print(f"After dedup: {len(merged)} products")

# ============================================================
# STEP 6: Save product_data_merged.json
# ============================================================

OUTPUT_JSON = f'{BASE_DIR}/product_data_merged.json'
with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(merged, f, ensure_ascii=False, indent=2)
print(f"JSON saved: {OUTPUT_JSON}")

# ============================================================
# Read Sheet2 A1 for footer note (used in both data.json and HTML)
footer_note = '产品数据会持续更新'
try:
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_FILLED, data_only=True)
    ws = wb['说明'] if '说明' in wb.sheetnames else wb.active
    a1_val = ws['A1'].value
    if a1_val:
        footer_note = str(a1_val).strip()
    wb.close()
except Exception:
    pass
# Fallback to primary Excel if filled version doesn't have the sheet
if footer_note == '产品数据会持续更新':
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PRIMARY, data_only=True)
        ws = wb['说明'] if '说明' in wb.sheetnames else wb.active
        a1_val = ws['A1'].value
        if a1_val:
            footer_note = str(a1_val).strip()
        wb.close()
    except Exception:
        pass

# Read "版本更新记录" sheet for app update info and full version history
app_update = {}
version_history = []
try:
    wb = openpyxl.load_workbook(EXCEL_FILLED, data_only=True)
    if '版本更新记录' in wb.sheetnames:
        ws = wb['版本更新记录']
        if ws.max_row >= 2:
            # Read all rows (skip header) for version history
            for row in range(2, ws.max_row + 1):
                ver = str(ws.cell(row, 1).value or '').strip()
                notes = str(ws.cell(row, 2).value or '').strip()
                if ver:
                    version_history.append({
                        'version': ver,
                        'notes': notes.replace('\\n', '\n') if notes else ''
                    })
            # Read the last row (newest version) for appUpdate
            last_row = ws.max_row
            ver = str(ws.cell(last_row, 1).value or '').strip()
            notes = str(ws.cell(last_row, 2).value or '').strip()
            force = str(ws.cell(last_row, 3).value or '').strip().upper() == 'TRUE'
            if ver:
                app_update = {
                    'latestVersion': ver,
                    'updateNotes': notes.replace('\\n', '\n'),
                    'apkUrl': f'https://kaiz1995.github.io/Oil-Matching-Search/RunmeiMatching.apk',
                    'forceUpdate': force,
                }
    wb.close()
except Exception:
    pass

# Read "发布页" sheet for release page URLs
release_pages = {}
try:
    wb = openpyxl.load_workbook(EXCEL_FILLED, data_only=True)
    if '发布页' in wb.sheetnames:
        ws = wb['发布页']
        for row in range(1, ws.max_row + 1):
            name = str(ws.cell(row, 1).value or '').strip()
            url = str(ws.cell(row, 2).value or '').strip()
            if name and url:
                release_pages[name] = url
    wb.close()
except Exception:
    pass

# STEP 6b: Save data.json (standalone, with version, for Android app)
# ============================================================

from datetime import date
today_str = date.today().isoformat()
data_export = {
    'version': f'{today_str}-v1',
    'updatedAt': today_str,
    'versionNote': footer_note,
    'totalProducts': len(merged),
    'products': merged,
}
if runmei_catalog:
    data_export['runmeiCatalog'] = runmei_catalog
if runmei_param_config:
    data_export['paramConfig'] = runmei_param_config
if getattr(args, 'release', False) and app_update:
    data_export['appUpdate'] = app_update
if version_history:
    data_export['versionHistory'] = version_history
if release_pages:
    data_export['releasePages'] = release_pages
OUTPUT_DATA_JSON = f'{BASE_DIR}/data.json'
with open(OUTPUT_DATA_JSON, 'w', encoding='utf-8') as f:
    json.dump(data_export, f, ensure_ascii=False, indent=2)
print(f"Data JSON saved: {OUTPUT_DATA_JSON} ({len(merged)} products)")

embedded_data_json = json.dumps(data_export, ensure_ascii=False, separators=(',', ':')).replace('</script>', '<\\/script>')

# ============================================================
# STEP 7: Generate HTML tool — update existing file in-place
# ============================================================

products = merged
# Normalize brand display names (prefer short Chinese-only form)
BRAND_PREFER = {
    '壳牌/Shell': '壳牌', '壳牌': '壳牌',
    '美孚/Mobil': '美孚', '美孚': '美孚',
    '雪佛龙/Chevron': '雪佛龙', '雪佛龙': '雪佛龙',
    '长城/Sinopec': '长城', '长城': '长城',
    '克鲁勃/Klüber': '克鲁勃', '克鲁勃': '克鲁勃',
}
for p in products:
    if p['brand'] in BRAND_PREFER:
        p['brand'] = BRAND_PREFER[p['brand']]

brands = sorted(set(p['brand'] for p in products))
categories = sorted(set(p['category'] for p in products))
runmei_series = sorted(set(p['series'] for p in runmei_catalog)) if runmei_catalog else []
brands_json = json.dumps(brands, ensure_ascii=False)
categories_json = json.dumps(categories, ensure_ascii=False)
cat_map_json = json.dumps(CAT_MAP, ensure_ascii=False)
runmei_series_json = json.dumps(runmei_series, ensure_ascii=False)
embedded_data_json = json.dumps(data_export, ensure_ascii=False, separators=(',', ':')).replace('</script>', '<\\/script>')

# Embed logo as base64
LOGO_PATH = f'{BASE_DIR}/logo3.png'
try:
    from PIL import Image
    import io
    img = Image.open(LOGO_PATH)
    aspect = img.width / img.height
    new_w = int(34 * aspect)
    img_resized = img.resize((new_w, 34), Image.LANCZOS)
    buf = io.BytesIO()
    img_resized.save(buf, format='PNG')
    logo_b64 = base64.b64encode(buf.getvalue()).decode()
except Exception as e:
    print(f'  [WARN] Cannot embed logo: {e}')
    logo_b64 = ''

# Read existing HTML, update data sections in-place
OUTPUT_HTML = f'{BASE_DIR}/RunmeiMatching-1.1.5.html'
if os.path.exists(OUTPUT_HTML):
    with open(OUTPUT_HTML, 'r', encoding='utf-8') as f:
        html = f.read()
else:
    # Fallback: create minimal HTML with essential structure
    html = '''<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>润美润滑油产品匹配工具</title></head><body></body></html>'''

import re

runmei_count = sum(1 for p in products if p['runmei'])
runmei_option_html = ''.join(f'<option value="{s}">{s}</option>' for s in runmei_series)

# 1. Update version note (bottom-left) — fallback for EMBEDDED_DATA
html = re.sub(
    r'<div class="version-note"[^>]*>.*?</div>',
    f'<div class="version-note" id="version-note">{footer_note}</div>',
    html, count=1, flags=re.DOTALL
)

# 2. Update header stats
html = re.sub(
    r'<span>竞品:.*?</span>',
    f'<span>竞品: {len(products)} 款</span>',
    html, count=1
)
html = re.sub(
    r'<span>润美:.*?</span>',
    f'<span>润美: {runmei_count} 款</span>',
    html, count=1
)
html = re.sub(
    r'<span>品类:.*?</span>',
    f'<span>品类: {len(categories)}</span>',
    html, count=1
)
html = re.sub(
    r'<span>品牌:.*?</span>',
    f'<span>品牌: {len(brands)}</span>',
    html, count=1
)

# 3. Update series dropdown options
html = re.sub(
    r'(<option value="">全部系列</option>).*?(?=</select>)',
    f'<option value="">全部系列</option>{runmei_option_html}',
    html, count=1, flags=re.DOTALL
)

# 4. Update JS variables
html = re.sub(
    r'var BRANDS\s*=\s*\[.*?\];',
    f'var BRANDS = {brands_json};',
    html, count=1
)
html = re.sub(
    r'var CATEGORIES\s*=\s*\[.*?\];',
    f'var CATEGORIES = {categories_json};',
    html, count=1
)
html = re.sub(
    r'var CAT_MAP\s*=\s*\{.*?\};',
    f'var CAT_MAP = {cat_map_json};',
    html, count=1, flags=re.DOTALL
)
html = re.sub(
    r'var RUNMEI_SERIES\s*=\s*\[.*?\];',
    f'var RUNMEI_SERIES = {runmei_series_json};',
    html, count=1
)
# 4. Update EMBEDDED_DATA with proper brace matching
emb_start_tag = 'var EMBEDDED_DATA = '
emb_start_idx = html.find(emb_start_tag)
if emb_start_idx >= 0:
    search_start = emb_start_idx + len(emb_start_tag)
    brace_depth = 0
    in_string = False
    esc = False
    end_idx = search_start
    for i in range(search_start, len(html)):
        ch = html[i]
        if esc:
            esc = False
            continue
        if ch == '\\' and in_string:
            esc = True
            continue
        if ch == '"':
            in_string = not in_string
            continue
        if not in_string:
            if ch == '{':
                brace_depth += 1
            elif ch == '}':
                brace_depth -= 1
                if brace_depth == 0:
                    end_idx = i + 1
                    break
    # Skip any old semicolons/newlines after closing brace
    skip_end = end_idx
    while skip_end < len(html) and html[skip_end] in ';\n\r ':
        skip_end += 1
    html = html[:emb_start_idx] + f'var EMBEDDED_DATA = {embedded_data_json};' + html[skip_end:]
else:
    print('  [WARN] Could not find var EMBEDDED_DATA in HTML')
# 7. Update logo
if logo_b64:
    html = re.sub(
        r'src="data:image/png;base64,[^"]*"',
        f'src="data:image/png;base64,{logo_b64}"',
        html, count=1
    )

with open(OUTPUT_HTML, 'w', encoding='utf-8') as f:
    f.write(html)

# Also update RunmeiMatching.html (latest-working copy)
LATEST_HTML = f'{BASE_DIR}/RunmeiMatching.html'
with open(LATEST_HTML, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"HTML tool saved: {OUTPUT_HTML}")
print(f"Latest HTML:    {LATEST_HTML}")
print(f"Size: {len(html.encode('utf-8')) / 1024:.1f} KB")
print(f"Products: {len(products)}")
print(f"With Runmei: {runmei_count}")
print(f"Brands: {len(brands)}")
print(f"Categories: {len(categories)}")
# ============================================================
# STEP 8: Release mode — build APK + create GitHub Release
# ============================================================

GITHUB_REPO = r'D:\HuaweiMoveData\Users\张大脸小太阳\Documents\cc workspace\Oil-Matching-Search'

if args.release and app_update:
    release_ver = app_update['latestVersion']
    ANDROID_PROJECT = r'C:\Users\张大脸小太阳\AndroidStudioProjects\RunmeiMatching'

    if not args.skip_apk:
        # Update versionName in build.gradle.kts
        gradle_file = f'{ANDROID_PROJECT}/app/build.gradle.kts'
        try:
            with open(gradle_file, 'r', encoding='utf-8') as f:
                content = f.read()
            content = re.sub(
                r'versionName\s*=\s*"[^"]*"',
                f'versionName = "{release_ver}"',
                content
            )
            with open(gradle_file, 'w', encoding='utf-8') as f:
                f.write(content)
            print(f'  versionName -> {release_ver} in build.gradle.kts')
        except Exception as e:
            print(f'  [WARN] Cannot update build.gradle.kts: {e}')

        # Build APK
        try:
            gradlew = os.path.join(ANDROID_PROJECT, 'gradlew.bat')
            r = subprocess.run([gradlew, 'assembleDebug'], capture_output=True, text=True, cwd=ANDROID_PROJECT)
            if r.returncode != 0:
                print(f'  [ERROR] APK build failed:\n{r.stderr.strip()[-500:]}')
            else:
                print(f'  APK build successful')
                apk_src = f'{ANDROID_PROJECT}/app/build/outputs/apk/debug/app-debug.apk'
                versioned_apk = f'RunmeiMatching-{release_ver}.apk'
                shutil.copy(apk_src, f'{GITHUB_REPO}/{versioned_apk}')
                shutil.copy(apk_src, f'{GITHUB_REPO}/RunmeiMatching.apk')
                print(f'  APK copied: {versioned_apk} + RunmeiMatching.apk')
        except Exception as e:
            print(f'  [ERROR] APK build: {e}')

    # Create GitHub Release
    tag = f'v{release_ver}'
    try:
        r = subprocess.run(['gh', 'release', 'view', tag, '--json', 'tagName',
            '--repo', 'kaiz1995/Oil-Matching-Search'], capture_output=True, text=True)
        if r.returncode == 0:
            print(f'  Release {tag} already exists, skipping')
        else:
            notes = app_update.get('updateNotes', '')
            versioned_apk = f'RunmeiMatching-{release_ver}.apk'
            r = subprocess.run([
                'gh', 'release', 'create', tag,
                f'{GITHUB_REPO}/{versioned_apk}',
                '--repo', 'kaiz1995/Oil-Matching-Search',
                '--title', tag, '--notes', notes,
            ], capture_output=True, text=True)
            if r.returncode == 0:
                print(f'  Release {tag} created')
            else:
                print(f'  [ERROR] Release create:\n{r.stderr.strip()[-500:]}')
    except FileNotFoundError:
        print(f'  [ERROR] gh CLI not found. Install it or create release manually:')
        print(f'    https://github.com/kaiz1995/Oil-Matching-Search/releases/new?tag={tag}')
        print(f'    Then upload: {GITHUB_REPO}/RunmeiMatching-{release_ver}.apk')
    except Exception as e:
        print(f'  [ERROR] Release: {e}')

# ============================================================
# STEP 9: Auto-deploy — copy to GitHub repo & Android project
# ============================================================

ANDROID_ASSETS = r'C:\Users\张大脸小太阳\AndroidStudioProjects\RunmeiMatching\app\src\main\assets'

deployed = []

# 1. Copy data.json to GitHub repo
try:
    shutil.copy(OUTPUT_DATA_JSON, f'{GITHUB_REPO}/data.json')
    deployed.append('GitHub 仓库 data.json')
except Exception as e:
    print(f'  [SKIP] GitHub 仓库: {e}')

# 1b. Copy pdfs/ to GitHub repo (for TDS/MSDS download)
try:
    PDFS_SRC = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pdfs')
    PDFS_DST = os.path.join(GITHUB_REPO, 'pdfs')
    os.makedirs(PDFS_DST, exist_ok=True)
    for fname in os.listdir(PDFS_SRC):
        if fname.endswith('.pdf'):
            src = os.path.join(PDFS_SRC, fname)
            dst = os.path.join(PDFS_DST, fname)
            if os.path.exists(dst):
                os.chmod(dst, 0o644)  # Remove read-only before overwrite
            shutil.copy2(src, dst)
            os.chmod(dst, 0o644)      # Ensure writable for git
    deployed.append('GitHub 仓库 pdfs/')
except Exception as e:
    print(f'  [SKIP] GitHub pdfs: {e}')

# 2. Git commit & push
if not args.skip_git:
    try:
        os.chdir(GITHUB_REPO)
        git_files = ['data.json', 'pdfs/']
        if os.path.exists(f'{GITHUB_REPO}/RunmeiMatching.apk'):
            git_files.append('RunmeiMatching.apk')
        import glob as _glob
        for _vf in _glob.glob(os.path.join(GITHUB_REPO, 'RunmeiMatching-*.apk')):
            _base = os.path.basename(_vf)
            if _base not in git_files:
                git_files.append(_base)
        subprocess.run(['git', 'add'] + git_files, capture_output=True, check=True)
        subprocess.run(['git', 'commit', '-m', f'更新润滑油数据 {today_str}'], capture_output=True)
        r = subprocess.run(['git', 'push'], capture_output=True, text=True)
        if r.returncode == 0:
            deployed.append('GitHub 已推送')
            # Auto-tag: create a unique data-YYYY-MM-DD-N tag for traceability
            try:
                existing = subprocess.run(['git', 'tag', '-l', f'data-{today_str}-*'],
                    capture_output=True, text=True).stdout.strip()
                count = len([t for t in existing.split('\n') if t]) + 1
                tag = f'data-{today_str}-{count}'
                subprocess.run(['git', 'tag', tag], capture_output=True, check=True)
                subprocess.run(['git', 'push', 'origin', tag], capture_output=True)
                print(f'  Tag: {tag}')
            except Exception as te:
                print(f'  [SKIP] Tag: {te}')
        else:
            print(f'  [SKIP] git push 失败: {r.stderr.strip() or r.stdout.strip()}')
    except Exception as e:
        print(f'  [SKIP] git 操作: {e}')

# 3. Copy data.json to Android fallback.json
try:
    shutil.copy(OUTPUT_DATA_JSON, f'{ANDROID_ASSETS}/fallback.json')
    deployed.append('Android fallback.json')
except Exception as e:
    print(f'  [SKIP] Android fallback: {e}')

# 4. Copy HTML to Android index.html
try:
    shutil.copy(OUTPUT_HTML, f'{ANDROID_ASSETS}/index.html')
    deployed.append('Android index.html')
except Exception as e:
    print(f'  [SKIP] Android index: {e}')

print(f"\nAuto-deploy: {', '.join(deployed) if deployed else '无'}")
print("\nDone. Pipeline complete.")
